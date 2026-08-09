"""
Plan B Scanner - 每日候选报告工作流
====================================
输入: unified_scanner.py 的扫描结果 CSV
步骤:
  1. 过滤 MCDX 双重确认信号 (resonance >= 55, red_ratio 0.60-0.85 sweet spot)
  2. 计算安全边际 = (DCF内在价值 - 现价) / DCF内在价值
  3. 标记入场候选 (现价 <= 85% IV，且 banker_flag 非 pure-hot-money)
  4. 生成 Excel 报告 (公式驱动，非硬编码)

⚠️ 说明：本脚本最初用模拟数据 (mock_scan_results.csv) 演示跑通，
输入CSV需要包含以下列（对应 unified_scanner.py + valuation_module.py 的真实输出）：
  ticker, name, sector, mcdx_resonance, red_ratio, banker_flag,
  current_price, dcf_intrinsic_value
若真实扫描结果的列名不同，需要先对齐或改这里的列名引用。

用法:
  python workflow.py <input_csv> <output_xlsx>
"""

import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def load_and_filter(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)

    # Plan B 双重确认信号: MCDX resonance >= 55 且 red_ratio 落在 0.60-0.85 sweet spot
    signal_mask = (df["mcdx_resonance"] >= 55) & (df["red_ratio"].between(0.60, 0.85))
    df = df[signal_mask].copy()

    # 按信号强度排序 (强 -> 弱)
    df = df.sort_values("mcdx_resonance", ascending=False).reset_index(drop=True)
    return df


def build_report(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Candidates"

    headers = [
        "Ticker", "Name", "Sector", "MCDX Resonance", "Red Ratio",
        "Banker Flag", "Current Price (RM)", "DCF Intrinsic Value (RM)",
        "Safety Margin (%)", "Entry Candidate?"
    ]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写入数据行 + 公式 (安全边际和入场判断用 Excel 公式计算，不用 Python 算好的硬数字)
    for i, row in df.iterrows():
        r = i + 2  # Excel 行号 (第1行是表头)
        ws.cell(row=r, column=1, value=row["ticker"])
        ws.cell(row=r, column=2, value=row["name"])
        ws.cell(row=r, column=3, value=row["sector"])
        ws.cell(row=r, column=4, value=row["mcdx_resonance"])
        ws.cell(row=r, column=5, value=row["red_ratio"])
        ws.cell(row=r, column=6, value=row["banker_flag"])
        ws.cell(row=r, column=7, value=row["current_price"])
        ws.cell(row=r, column=8, value=row["dcf_intrinsic_value"])
        # 安全边际 = (IV - 现价) / IV，公式引用本行的 G(现价) 和 H(IV)
        ws.cell(row=r, column=9, value=f"=(H{r}-G{r})/H{r}")
        # 入场候选: 安全边际 >= 15% (现价 <= 85% IV) 且不是纯游资
        ws.cell(row=r, column=10,
                 value=f'=IF(AND(I{r}>=0.15,F{r}<>"pure-hot-money"),"YES","WATCH")')

    # 数字格式
    for r in range(2, len(df) + 2):
        ws.cell(row=r, column=7).number_format = '"RM"#,##0.00'
        ws.cell(row=r, column=8).number_format = '"RM"#,##0.00'
        ws.cell(row=r, column=9).number_format = '0.0%'
        ws.cell(row=r, column=5).number_format = '0.00'

    # 条件格式: 安全边际列 (I列) 绿色=达标, 红色=不足
    last_row = len(df) + 1
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        f"I2:I{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.15"], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"I2:I{last_row}",
        CellIsRule(operator="lessThan", formula=["0.15"], fill=red_fill)
    )

    # 列宽
    widths = [10, 16, 12, 15, 10, 18, 16, 20, 15, 15]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 备注 / 假设说明
    note_row = last_row + 2
    ws.cell(row=note_row, column=1,
            value="说明: MCDX sweet spot = resonance>=55 & red_ratio 0.60-0.85 (511信号回测验证, ~62%胜率, 60日持有期)")
    ws.cell(row=note_row, column=1).font = Font(name="Arial", italic=True, size=9)
    ws.cell(row=note_row + 1, column=1,
            value="入场判定: 安全边际>=15% (现价<=85% DCF内在价值) 且非纯游资 (pure-hot-money) 类型")
    ws.cell(row=note_row + 1, column=1).font = Font(name="Arial", italic=True, size=9)

    wb.save(output_path)


if __name__ == "__main__":
    input_csv = sys.argv[1] if len(sys.argv) > 1 else "mock_scan_results.csv"
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "daily_candidates.xlsx"

    filtered = load_and_filter(input_csv)
    print(f"扫描输入: {input_csv}")
    print(f"筛选出 {len(filtered)} 只符合双重确认信号的股票:")
    print(filtered[["ticker", "name", "mcdx_resonance", "red_ratio", "banker_flag"]].to_string(index=False))

    build_report(filtered, output_xlsx)
    print(f"\n报告已生成: {output_xlsx}")
